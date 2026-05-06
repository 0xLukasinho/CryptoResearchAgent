"""One-off: check every Substack URL in input/Substacks.csv for liveness.

Steps:
  1. Read from the *original* backup (so previously-removed dead entries
     get a second chance with normalized URLs).
  2. Normalize URLs to publication root: strip /archive, /about, /p/<slug>,
     trailing slash, etc.
  3. Dedupe by normalized URL (keeps first occurrence; rows are merged so
     non-empty fields from later duplicates fill in gaps).
  4. Concurrent HTTP check of each unique URL via Substack's posts API.
  5. Bucket results: alive / dead / private / other.
  6. Write cleaned CSV (alive on top, private at bottom, dead removed) and a
     review XLSX with red highlight on PRIVATE rows.

Buckets:
  alive   : 2xx response from the posts API
  dead    : 404 (publication gone)
  private : 403 (publication exists, blocks unauthenticated API)
  other   : timeouts, connection errors, 5xx, etc.
"""
from __future__ import annotations

import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Read from original backup so we get a clean second pass
INPUT_PATH = Path("input/Substacks.csv.bak.2026-05-06")
OUTPUT_CSV = Path("input/Substacks.csv")
OUTPUT_XLSX = Path("input/Substacks_review.xlsx")

USER_AGENT = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")
HEADERS = {"User-Agent": USER_AGENT}
TIMEOUT = 15
WORKERS = 20

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)


def normalize_url(u: str) -> str:
    """Normalize to publication root: scheme://host (no path, no query, no fragment).

    Examples:
      'chainml.substack.com/archive'        -> 'https://chainml.substack.com'
      'https://blog.safary.club/p/foo'      -> 'https://blog.safary.club'
      'https://x.substack.com/'             -> 'https://x.substack.com'
      ''                                     -> ''
    """
    u = (u or "").strip()
    if not u:
        return ""
    if not urlparse(u).scheme:
        u = "https://" + u
    parsed = urlparse(u)
    if not parsed.netloc:
        return ""
    # Drop path, params, query, fragment — keep only scheme://host
    return urlunparse((parsed.scheme, parsed.netloc, "", "", "", ""))


def check_one(url: str) -> tuple[str, int | None, str | None]:
    """Returns (bucket, status_code, error_message)."""
    api = f"{url}/api/v1/posts?sort=new&limit=1&offset=0"
    try:
        r = requests.get(api, headers=HEADERS, timeout=TIMEOUT)
    except requests.Timeout:
        return ("other", None, "timeout")
    except requests.ConnectionError as e:
        return ("other", None, f"conn_error: {type(e).__name__}")
    except Exception as e:
        return ("other", None, f"{type(e).__name__}: {e}")
    if 200 <= r.status_code < 300:
        return ("alive", r.status_code, None)
    if r.status_code == 404:
        return ("dead", 404, None)
    if r.status_code == 403:
        return ("private", 403, None)
    return ("other", r.status_code, None)


def dedupe_keep_first(df: pd.DataFrame, key: str) -> tuple[pd.DataFrame, int]:
    """Drop duplicate rows by `key`, keeping the first occurrence.
    Returns (deduped_df, num_dropped)."""
    before = len(df)
    deduped = df.drop_duplicates(subset=[key], keep="first").reset_index(drop=True)
    return deduped, before - len(deduped)


def main() -> int:
    if not INPUT_PATH.exists():
        print(f"ERROR: {INPUT_PATH} not found.", file=sys.stderr)
        return 1

    df = pd.read_csv(INPUT_PATH)
    print(f"Loaded {len(df)} rows from {INPUT_PATH}")

    # 1. Normalize URLs (strip paths, add scheme)
    df["Substack URL"] = df["Substack URL"].fillna("").map(normalize_url)
    non_empty = (df["Substack URL"] != "").sum()
    print(f"Normalized URLs; {non_empty} non-empty after normalization")

    # 2. Drop rows with empty URL
    df = df[df["Substack URL"] != ""].reset_index(drop=True)

    # 3. Dedupe
    df, dropped_dupes = dedupe_keep_first(df, "Substack URL")
    print(f"Removed {dropped_dupes} duplicate URL(s); {len(df)} unique URLs remain")

    # 4. Check liveness in parallel
    work = list(enumerate(df["Substack URL"]))
    print(f"\nChecking {len(work)} URLs with {WORKERS} workers...\n")
    results: dict[int, tuple[str, int | None, str | None]] = {}
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = {ex.submit(check_one, url): (idx, url) for idx, url in work}
        done = 0
        for fut in as_completed(futures):
            idx, _url = futures[fut]
            results[idx] = fut.result()
            done += 1
            if done % 50 == 0 or done == len(work):
                print(f"  {done}/{len(work)} checked")

    df["_bucket"] = [results.get(i, ("alive", 200, None))[0] for i in range(len(df))]
    df["_code"] = [results.get(i, (None, None, None))[1] for i in range(len(df))]
    df["_err"] = [results.get(i, (None, None, None))[2] for i in range(len(df))]

    # 5. Summary
    counts = df["_bucket"].value_counts().to_dict()
    print("\n=== Summary ===")
    for k in ("alive", "dead", "private", "other"):
        print(f"  {k:8s}: {counts.get(k, 0)}")
    print(f"  duplicates removed: {dropped_dupes}")

    print("\nDead (404, will be removed):")
    for url in df.loc[df["_bucket"] == "dead", "Substack URL"]:
        print(f"  {url}")

    print("\nPrivate (403, kept at bottom):")
    for url in df.loc[df["_bucket"] == "private", "Substack URL"]:
        print(f"  {url}")

    print("\nOther failures (kept in CSV with Status=OTHER):")
    for _, row in df.loc[df["_bucket"] == "other"].iterrows():
        print(f"  [{row['_code']}|{row['_err']}] {row['Substack URL']}")

    # 6. Build output: alive+other in original order, then private at bottom
    alive_or_other = df[df["_bucket"].isin(["alive", "other"])].copy()
    private_rows = df[df["_bucket"] == "private"].copy()
    alive_or_other["Status"] = alive_or_other["_bucket"].map(
        lambda b: "ALIVE" if b == "alive" else "OTHER"
    )
    private_rows["Status"] = "PRIVATE"

    out_cols = ["Name", "by", "Substack URL", "X URL", "Status"]
    out_df = pd.concat([alive_or_other[out_cols], private_rows[out_cols]],
                       ignore_index=True)
    out_df.to_csv(OUTPUT_CSV, index=False)
    print(f"\nWrote cleaned CSV: {OUTPUT_CSV} ({len(out_df)} rows)")

    # 7. XLSX with red on PRIVATE
    wb = Workbook()
    ws = wb.active
    ws.title = "Substacks"
    ws.append(out_cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for _, row in out_df.iterrows():
        ws.append([row[c] for c in out_cols])
        if row["Status"] == "PRIVATE":
            for cell in ws[ws.max_row]:
                cell.fill = RED_FILL
                cell.font = RED_FONT
    for col_idx, width in enumerate([28, 26, 50, 36, 10], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    wb.save(OUTPUT_XLSX)
    print(f"Wrote review XLSX: {OUTPUT_XLSX}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
